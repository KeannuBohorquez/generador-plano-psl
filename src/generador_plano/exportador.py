# =============================================================
#  exportador.py — Exportacion del archivo plano a Excel
# =============================================================

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import HEADERS

# Anchos de columna en caracteres (mismo orden que HEADERS)
_ANCHOS = [10, 10, 6, 8, 8, 14, 22, 22, 10, 12, 8, 10, 14, 10, 16, 12, 12, 12, 12, 50]

# Formato de celda por columna (indice 1-based).
# Columnas marcadas como texto ("@") reciben ademas conversion a str
# para evitar la advertencia "numero almacenado como texto" en Excel.
#
# Mapa: col_idx -> (number_format, forzar_texto)
#   number_format : cadena de formato de Excel
#   forzar_texto  : si True, el valor se convierte a str antes de escribir
#
# Orden de columnas (del constructor.py):
#  1 compania | 2 division | 3 ano | 4 periodo | 5 fuente
#  6 N comprobante | 7 fecha contab. | 8 fecha transac. | 9 operacion
# 10 cod cuenta | 11 centro | 12 concepto | 13 Tercero | 14 Documento
# 15 vlr moneda lc | 16 cod mond extr | 17 valor moneda | 18 valor base
# 19 origen | 20 comentario
_COL_FMT: dict[int, tuple[str, bool]] = {
    1:  ("General", False),  # compania
    2:  ("General", False),  # division
    3:  ("General", False),  # ano
    4:  ("General", False),  # periodo
    5:  ("@",       True ),  # fuente            → texto
    6:  ("@",       True ),  # N comprobante     → texto
    7:  ("@",       True ),  # fecha contab.     → texto dd/mm/yyyy
    8:  ("@",       True ),  # fecha transac.    → texto dd/mm/yyyy
    9:  ("@",       True ),  # operacion         → texto
    10: ("@",       True ),  # cod cuenta        → texto
    11: ("General", False),  # centro            → general
    12: ("General", False),  # concepto          → general
    13: ("@",       True ),  # Tercero           → texto
    14: ("@",       True ),  # Documento         → texto
    15: ("#,##0.00",False),  # vlr moneda lc     → Numero
    16: ("@",       True ),  # cod mond extr     → texto
    17: ("General", False),  # valor moneda      → general
    18: ("General", False),  # valor base        → general
    19: ("@",       True ),  # origen            → texto
    20: ("General", False),  # comentario        → general
}


def _str_seguro(val) -> str:
    """Convierte un valor a str, devolviendo '' si es None."""
    if val is None:
        return ""
    return str(val)


def exportar_excel(
    filas: list[list],
    ruta_salida: str,
    indices_faltante: list[int] | None = None,
) -> None:
    """
    Escribe las filas del archivo plano en un Excel con formato.

    Formatos de columna aplicados segun requerimiento:
    - Texto (@)  : fuente, N comprobante, fechas, operacion, cod cuenta,
                   Tercero, Documento, cod mond extr, origen
    - Numero     : vlr moneda lc  (#,##0.00)
    - General    : compania, division, ano, periodo, centro, concepto,
                   valor moneda, valor base, comentario

    Args:
        filas: Lista de filas generadas por construir_filas().
        ruta_salida: Ruta completa del archivo .xlsx a crear.
        indices_faltante: Indices base-0 de filas que representan
                          dinero faltante (se resaltan en rojo).

    Raises:
        PermissionError: Si el archivo esta abierto en Excel.
    """
    indices_faltante = set(indices_faltante or [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Archivo Plano"

    # ── Encabezado ────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ── Datos con filas alternas ──────────────────────────────
    alt_fill      = PatternFill("solid", fgColor="DCE6F1")
    faltante_fill = PatternFill("solid", fgColor="FF0000")
    faltante_font = Font(color="FFFFFF", bold=True, size=10)

    for ri, fila in enumerate(filas, 2):
        idx_base0   = ri - 2
        es_faltante = idx_base0 in indices_faltante

        if es_faltante:
            fill = faltante_fill
        elif ri % 2 == 0:
            fill = alt_fill
        else:
            fill = None

        for ci, val in enumerate(fila, 1):
            fmt, forzar_texto = _COL_FMT.get(ci, ("General", False))

            # Convertir a texto antes de escribir si la columna lo requiere
            if forzar_texto:
                val = _str_seguro(val)

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.number_format = fmt

            if fill:
                cell.fill = fill
            if es_faltante:
                cell.font = faltante_font

    # ── Anchos y panel fijo ───────────────────────────────────
    for ci, ancho in enumerate(_ANCHOS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = ancho
    ws.freeze_panes = "A2"

    wb.save(ruta_salida)
